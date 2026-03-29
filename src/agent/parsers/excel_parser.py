"""Excel parser for .xlsx and .xls files.

This parser handles Excel files (.xlsx, .xls), extracting:
- Sheet names
- Cell values
- Table structure
- CSV-like output for easy chunking
"""

from dataclasses import dataclass
from pathlib import Path
from typing import List, Optional

from src.agent.schemas.document import (
    DocumentPage,
    DocumentSection,
    ParsedDocument,
)
from src.agent.parsers.base import BaseParser, ParseError

# Try to import openpyxl
try:
    from openpyxl import load_workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# Try to import xlrd for .xls files
try:
    import xlrd
    XLRD_AVAILABLE = True
except ImportError:
    XLRD_AVAILABLE = False


@dataclass
class ExcelParserConfig:
    """Configuration for Excel parser."""
    include_formulas: bool = False  # Whether to include formula text
    include_sheet_names: bool = True  # Include sheet name as section title
    max_rows_per_section: int = 100  # Split large sheets into sections


class ExcelParser(BaseParser):
    """Parser for Excel (.xlsx, .xls) files.
    
    This parser:
    1. Reads all sheets from Excel file
    2. Extracts cell values
    3. Preserves table structure
    4. Converts to readable text format
    
    Attributes:
        config: Parser configuration options
    """
    
    supported_extensions = [".xlsx", ".xls"]
    
    def __init__(self, config: Optional[ExcelParserConfig] = None):
        """Initialize Excel parser.
        
        Args:
            config: Optional configuration. Uses defaults if not provided.
        """
        self.config = config or ExcelParserConfig()
    
    def parse(self, file_path: str | Path) -> ParsedDocument:
        """Parse an Excel file.
        
        Args:
            file_path: Path to the Excel file.
            
        Returns:
            ParsedDocument with parsed content.
            
        Raises:
            ParseError: If parsing fails.
        """
        if not OPENPYXL_AVAILABLE:
            raise ParseError(
                "openpyxl is not installed. "
                "Install it with: pip install openpyxl"
            )
        
        path = self.validate_file(file_path)
        ext = path.suffix.lower()
        
        try:
            if ext == ".xlsx":
                return self._parse_xlsx(path)
            elif ext == ".xls":
                return self._parse_xls(path)
            else:
                raise ParseError(f"Unsupported Excel format: {ext}")
        except ParseError:
            raise
        except Exception as e:
            raise ParseError(f"Failed to parse Excel file: {e}") from e
    
    def _parse_xlsx(self, path: Path) -> ParsedDocument:
        """Parse .xlsx file using openpyxl.
        
        Args:
            path: File path
            
        Returns:
            ParsedDocument with parsed content
        """
        try:
            wb = load_workbook(filename=str(path), data_only=True)
        except Exception as e:
            raise ParseError(f"Failed to open .xlsx file: {e}") from e
        
        sections: List[DocumentSection] = []
        char_pos = 0
        
        sheet_names = wb.sheetnames
        
        for sheet_idx, sheet_name in enumerate(sheet_names):
            ws = wb[sheet_name]
            
            # Get sheet title
            if self.config.include_sheet_names:
                sheet_title = f"Sheet: {sheet_name}"
            else:
                sheet_title = None
            
            # Extract data from worksheet
            sheet_text, row_count = self._extract_sheet_text(ws)
            
            # Create section for this sheet
            if sheet_text.strip():
                sections.append(DocumentSection(
                    title=sheet_title,
                    level=1,
                    page_no=1,
                    char_start=char_pos,
                    char_end=char_pos + len(sheet_text),
                    text=sheet_text,
                ))
                char_pos += len(sheet_text) + 1
        
        wb.close()
        
        # If no sections, create empty document
        if not sections:
            sections.append(DocumentSection(
                title=None,
                level=1,
                page_no=1,
                char_start=0,
                char_end=0,
                text="",
            ))
        
        # Build full text
        full_text = "\n\n".join([s.text for s in sections])
        
        # Create single page
        page = DocumentPage(
            page_no=1,
            text=full_text,
            sections=sections,
        )
        
        return ParsedDocument(
            source_file=str(path),
            total_pages=1,
            pages=[page],
            sections=sections,
        )
    
    def _parse_xls(self, path: Path) -> ParsedDocument:
        """Parse .xls file using xlrd.
        
        Args:
            path: File path
            
        Returns:
            ParsedDocument with parsed content
        """
        if not XLRD_AVAILABLE:
            raise ParseError(
                "xlrd is not installed for .xls support. "
                "Install it with: pip install xlrd"
            )
        
        try:
            wb = xlrd.open_workbook(str(path))
        except Exception as e:
            raise ParseError(f"Failed to open .xls file: {e}") from e
        
        sections: List[DocumentSection] = []
        char_pos = 0
        
        sheet_names = wb.sheet_names()
        
        for sheet_name in sheet_names:
            ws = wb.sheet_by_name(sheet_name)
            
            # Get sheet title
            if self.config.include_sheet_names:
                sheet_title = f"Sheet: {sheet_name}"
            else:
                sheet_title = None
            
            # Extract data from worksheet
            sheet_text, row_count = self._extract_xls_sheet_text(ws)
            
            # Create section for this sheet
            if sheet_text.strip():
                sections.append(DocumentSection(
                    title=sheet_title,
                    level=1,
                    page_no=1,
                    char_start=char_pos,
                    char_end=char_pos + len(sheet_text),
                    text=sheet_text,
                ))
                char_pos += len(sheet_text) + 1
        
        # If no sections, create empty document
        if not sections:
            sections.append(DocumentSection(
                title=None,
                level=1,
                page_no=1,
                char_start=0,
                char_end=0,
                text="",
            ))
        
        # Build full text
        full_text = "\n\n".join([s.text for s in sections])
        
        # Create single page
        page = DocumentPage(
            page_no=1,
            text=full_text,
            sections=sections,
        )
        
        return ParsedDocument(
            source_file=str(path),
            total_pages=1,
            pages=[page],
            sections=sections,
        )
    
    def _extract_sheet_text(self, ws) -> tuple[str, int]:
        """Extract text from a worksheet.
        
        Args:
            ws: openpyxl Worksheet object
            
        Returns:
            Tuple of (extracted text, row count)
        """
        rows = []
        row_count = 0
        
        # Get max row and column
        max_row = ws.max_row
        max_col = ws.max_column
        
        if max_row == 0 or max_col == 0:
            return "", 0
        
        # Read row by row
        for row_idx in range(1, min(max_row + 1, self.config.max_rows_per_section + 1)):
            row_values = []
            for col_idx in range(1, max_col + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                value = cell.value
                
                if value is None:
                    row_values.append("")
                elif isinstance(value, str):
                    row_values.append(value)
                else:
                    row_values.append(str(value))
            
            # Only add row if it has content
            if any(v for v in row_values):
                rows.append(" | ".join(row_values))
                row_count += 1
        
        # If more rows exist, add indicator
        if max_row > self.config.max_rows_per_section:
            rows.append(f"... ({max_row - self.config.max_rows_per_section} more rows)")
        
        return "\n".join(rows), row_count
    
    def _extract_xls_sheet_text(self, ws) -> tuple[str, int]:
        """Extract text from an xlrd Worksheet.
        
        Args:
            ws: xlrd Worksheet object
            
        Returns:
            Tuple of (extracted text, row count)
        """
        rows = []
        row_count = 0
        
        # Get max row and column
        max_row = ws.nrows
        max_col = ws.ncols
        
        if max_row == 0 or max_col == 0:
            return "", 0
        
        # Read row by row
        for row_idx in range(min(max_row, self.config.max_rows_per_section)):
            row_values = []
            for col_idx in range(max_col):
                cell = ws.cell(rowx=row_idx, colx=col_idx)
                value = cell.value
                
                if value is None:
                    row_values.append("")
                elif isinstance(value, str):
                    row_values.append(value)
                else:
                    row_values.append(str(value))
            
            # Only add row if it has content
            if any(v for v in row_values):
                rows.append(" | ".join(row_values))
                row_count += 1
        
        # If more rows exist, add indicator
        if max_row > self.config.max_rows_per_section:
            rows.append(f"... ({max_row - self.config.max_rows_per_section} more rows)")
        
        return "\n".join(rows), row_count


__all__ = [
    "ExcelParser",
    "ExcelParserConfig",
]
